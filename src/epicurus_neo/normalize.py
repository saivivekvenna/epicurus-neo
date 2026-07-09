from __future__ import annotations

from pathlib import Path
from typing import Any
from zipfile import ZipFile

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from epicurus_neo.features import add_contrastive_features
from epicurus_neo.schema import add_normalized_columns, validate_schema


PEPTIDE_COLS = [
    "mutant_seq",
    "mut_seq",
    "mutant_peptide",
    "Mut Epitope",
    "Mutant Minimal Peptide",
    "Mutant Mmp",
    "Mutated Minimal Peptide",
    "Mutant Nmer",
    "mutant_nmer",
]

WILDTYPE_COLS = [
    "wt_seq",
    "wildtype_seq",
    "wildtype_peptide",
    "Wt Epitope",
    "Wildtype Minimal Peptide",
    "Wildtype Mmp",
    "Wildtype Nmer",
    "wt_nmer",
]

HLA_COLS = [
    "hla_allele",
    "HLA",
    "Allele",
    "HLA Allele",
    "Best Allele",
    "best_allele",
    "mutant_best_allele",
]

PATIENT_COLS = ["patient", "Patient", "patient_id", "Sample", "sample", "ID"]
STUDY_COLS = ["dataset", "Dataset", "study_id", "Study"]
GENE_COLS = ["gene", "Gene Name", "Gene", "gene_symbol"]


def _read_delimited(handle: Any, *, suffix: str) -> pd.DataFrame:
    if suffix in {".tsv", ".txt"}:
        frame = pd.read_csv(handle, sep="\t")
        if len(frame.columns) == 1 and "," in str(frame.columns[0]):
            if hasattr(handle, "seek"):
                handle.seek(0)
            return pd.read_csv(handle)
        return frame
    return pd.read_csv(handle)


def _read_xlsx_with_repaired_dimensions(path: Path) -> pd.DataFrame:
    """Read publisher workbooks that incorrectly declare their used range as A1."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    worksheet.reset_dimensions()
    rows = worksheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return pd.DataFrame()
    columns = [str(value) if value is not None else f"unnamed_{idx}" for idx, value in enumerate(header)]
    return pd.DataFrame(rows, columns=columns)


def read_table(path: str | Path, *, zip_member: str | None = None) -> pd.DataFrame:
    table_path = Path(path)
    suffix = table_path.suffix.lower()
    if suffix == ".zip":
        with ZipFile(table_path) as archive:
            names = [name for name in archive.namelist() if not name.endswith("/")]
            if zip_member is not None:
                if zip_member not in names:
                    raise ValueError(f"Zip member {zip_member!r} not found in {table_path}")
                names = [zip_member]
            elif len(names) != 1:
                raise ValueError(f"Expected one table file in zip archive, found {names}")
            with archive.open(names[0]) as handle:
                inner_suffix = Path(names[0]).suffix.lower()
                return _read_delimited(handle, suffix=inner_suffix)
    if suffix == ".csv":
        return pd.read_csv(table_path)
    if suffix in {".tsv", ".txt"}:
        return _read_delimited(table_path, suffix=suffix)
    if suffix == ".xlsx":
        return _read_xlsx_with_repaired_dimensions(table_path)
    raise ValueError(f"Unsupported input table format: {table_path}")


def _first_existing(frame: pd.DataFrame, candidates: list[str]) -> str | None:
    for column in candidates:
        if column in frame.columns:
            return column
    return None


def _series_or_default(frame: pd.DataFrame, candidates: list[str], default: Any = "") -> pd.Series:
    column = _first_existing(frame, candidates)
    if column is None:
        return pd.Series([default] * len(frame), index=frame.index)
    return frame[column]


def _label_from_response(value: Any) -> str:
    if value is None or pd.isna(value):
        return "unknown"
    normalized = str(value).strip().lower()
    if normalized in {"cd8", "positive", "pos", "reactive", "immunogenic", "1", "true", "yes"}:
        return "positive"
    if normalized in {"negative", "neg", "non-reactive", "nonreactive", "0", "false", "no"}:
        return "negative"
    if normalized in {"not_tested", "not tested", "unknown", "untested", "nan", ""}:
        return "unknown"
    return "unknown"


def _label_column(frame: pd.DataFrame) -> pd.Series:
    candidates = [
        "response_type",
        "Response Type",
        "response",
        "Response",
        "Screening Status",
        "immunogenicity",
        "Immunogenicity",
        "reactivity",
        "Reactivity",
        "TIL Reactivity",
    ]
    source = _series_or_default(frame, candidates, "unknown")
    return source.map(_label_from_response)


def _candidate_ids(source_dataset: str, frame: pd.DataFrame) -> pd.Series:
    if "candidate_id" in frame.columns:
        return frame["candidate_id"].astype(str)
    return pd.Series(
        [f"{source_dataset}:{idx}" for idx in range(len(frame))],
        index=frame.index,
    )


def _copy_numeric_features(source: pd.DataFrame, target: pd.DataFrame) -> pd.DataFrame:
    out = target.copy()
    for column in source.columns:
        if column in out.columns:
            continue
        numeric = pd.to_numeric(source[column], errors="coerce")
        if numeric.notna().any():
            out[column] = numeric
    return out


def normalize_candidate_table(
    frame: pd.DataFrame,
    *,
    source_dataset: str,
    study_default: str | None = None,
) -> pd.DataFrame:
    labels = _label_column(frame)
    study = _series_or_default(frame, STUDY_COLS, study_default or source_dataset)
    patient = _series_or_default(frame, PATIENT_COLS, "")

    out = pd.DataFrame(
        {
            "candidate_id": _candidate_ids(source_dataset, frame),
            "source_dataset": source_dataset,
            "study_id": study.astype(str),
            "patient_id": patient.astype(str),
            "hla_allele": _series_or_default(frame, HLA_COLS, "").astype(str),
            "mutant_peptide": _series_or_default(frame, PEPTIDE_COLS, "").astype(str),
            "wildtype_peptide": _series_or_default(frame, WILDTYPE_COLS, "").astype(str),
            "label": labels,
            "label_weight": np.where(labels == "unknown", 0.0, 1.0),
            "assay_type": _series_or_default(frame, ["assay_type", "Assay", "Assay Type"], "unknown"),
        }
    )

    gene_col = _first_existing(frame, GENE_COLS)
    if gene_col is not None:
        out["gene_symbol"] = frame[gene_col].astype(str)

    out = _copy_numeric_features(frame, out)
    out = add_contrastive_features(out)
    out = add_normalized_columns(out)
    report = validate_schema(out)
    if not report.ok:
        raise ValueError(f"Normalized table failed canonical schema validation: {report}")
    return out


def normalize_neoranking_neopep(path: str | Path) -> pd.DataFrame:
    frame = read_table(path)
    out = normalize_candidate_table(
        frame,
        source_dataset="neoranking",
        study_default="neoranking",
    )

    rename = {
        "rnaseq_TPM": "expression_tpm",
        "CCF": "clonality_ccf",
        "mutant_rank_netMHCpan": "netmhcpan_mutant_rank",
        "mutant_rank_PRIME": "prime_mutant_rank",
        "mut_Rank_Stab": "netmhcstab_mutant_rank",
        "mut_binding_score": "binding_score",
        "DAI_NetMHC": "dai_netmhc",
        "DAI_MixMHC": "dai_mixmhc",
        "bestWTMatchScore_I": "self_similarity_score",
        "bestMutationScore_I": "foreignness_mutation_score",
    }
    for old, new in rename.items():
        if old in out.columns and new not in out.columns:
            out[new] = out[old]
    return out


def normalize_gartner_table(path: str | Path) -> pd.DataFrame:
    frame = read_table(path)
    out = normalize_candidate_table(
        frame,
        source_dataset="gartner_nci",
        study_default="gartner_nci",
    )
    if "Screening Status" in frame.columns:
        status = frame["Screening Status"].astype(str).str.strip().str.lower()
        labels = status.map(
            {
                "cd8": "positive",
                "1": "positive",
                "0": "negative",
                "-": "negative",
                "unscreened": "unknown",
            }
        ).fillna("unknown")
        out["label"] = labels
        out["label_weight"] = np.where(labels == "unknown", 0.0, 1.0)

    rename = {
        "Rank NetMHC": "baseline_netmhc_rank",
        "Rank Nmer Model": "baseline_gartner_nmer_rank",
        "Rank Mmp Model": "baseline_gartner_mmp_rank",
        "AUC NetMHC": "source_auc_netmhc",
        "AUC NMER Model": "source_auc_nmer_model",
    }
    for old, new in rename.items():
        if old in out.columns and new not in out.columns:
            out[new] = out[old]
    return out


def normalize_tesla_table(path: str | Path) -> pd.DataFrame:
    table_path = Path(path)
    if table_path.suffix.lower() in {".xlsx", ".xls"}:
        frame = pd.read_excel(table_path)
    else:
        frame = read_table(table_path)

    required = {"peptide", "target_value", "allele"}
    missing = required - set(frame.columns)
    if missing:
        raise ValueError(f"TESLA table missing required columns: {sorted(missing)}")

    labels = frame["target_value"].map(lambda value: "positive" if int(value) == 1 else "negative")
    out = pd.DataFrame(
        {
            "candidate_id": [f"tesla:{idx}" for idx in range(len(frame))],
            "source_dataset": "tesla",
            "study_id": "tesla_wells_2020",
            "patient_id": "tesla_unknown",
            "hla_allele": frame["allele"].astype(str),
            "mutant_peptide": frame["peptide"].astype(str),
            "wildtype_peptide": "",
            "label": labels,
            "label_weight": 1.0,
            "assay_type": "tesla_target_value",
        }
    )
    out = add_normalized_columns(out)
    report = validate_schema(out)
    if not report.ok:
        raise ValueError(f"Normalized TESLA table failed canonical schema validation: {report}")
    return out


def normalize_bigmhc_table(path: str | Path, *, zip_member: str | None = None) -> pd.DataFrame:
    frame = read_table(path, zip_member=zip_member)
    required = {"mhc", "pep", "tgt"}
    missing = required - set(frame.columns)
    if missing:
        raise ValueError(f"BigMHC table missing required columns: {sorted(missing)}")

    split_name = Path(zip_member or path).stem
    labels = frame["tgt"].map(lambda value: "positive" if int(value) == 1 else "negative")
    out = pd.DataFrame(
        {
            "candidate_id": [f"bigmhc:{split_name}:{idx}" for idx in range(len(frame))],
            "source_dataset": "bigmhc",
            "study_id": f"bigmhc_{split_name}",
            "patient_id": "bigmhc_" + frame["mhc"].astype(str),
            "hla_allele": frame["mhc"].astype(str),
            "mutant_peptide": frame["pep"].astype(str),
            "wildtype_peptide": frame["wtp"].astype(str) if "wtp" in frame.columns else "",
            "label": labels,
            "label_weight": 1.0,
            "assay_type": "bigmhc_immunogenicity",
        }
    )
    if "gene" in frame.columns:
        out["gene_symbol"] = frame["gene"].astype(str)

    out = _copy_numeric_features(frame, out)
    rename = {
        "BigMHC_EL": "bigmhc_el_score",
        "BigMHC_IM": "bigmhc_im_score",
        "BigMHC_ELIM": "bigmhc_elim_score",
        "NetMHCpan-4.1_Scores": "netmhcpan_41_score",
        "NetMHCpan-4.1_Ranks": "netmhcpan_41_rank",
        "MHCflurry-2.0_Scores": "mhcflurry_20_score",
        "MHCflurry-2.0_Ranks": "mhcflurry_20_rank",
        "PRIME-2.0_Scores": "prime_20_score",
        "PRIME-2.0_Ranks": "prime_20_rank",
    }
    for old, new in rename.items():
        if old in out.columns and new not in out.columns:
            out[new] = out[old]

    out = add_contrastive_features(out)
    out = add_normalized_columns(out)
    report = validate_schema(out)
    if not report.ok:
        raise ValueError(f"Normalized BigMHC table failed canonical schema validation: {report}")
    return out


def normalize_cd8_multimer_2025(path: str | Path) -> pd.DataFrame:
    """Normalize the 8,103-candidate patient pHLA multimer screen."""
    frame = read_table(path)
    required = {
        "Patient ID",
        "WT epitope",
        "MUT epitope",
        "HLA",
        "Response",
        "Dataset",
        "Tumor type",
    }
    missing = required - set(frame.columns)
    if missing:
        raise ValueError(f"CD8 multimer table missing required columns: {sorted(missing)}")

    frame = frame.copy()
    frame["_source_row"] = np.arange(2, len(frame) + 2)
    dedup_key = ["Patient ID", "MUT epitope", "HLA"]
    label_counts = frame.groupby(dedup_key, dropna=False)["Response"].nunique()
    conflicting = label_counts[label_counts > 1]
    if not conflicting.empty:
        raise ValueError(
            f"CD8 multimer table has {len(conflicting)} duplicate candidates with conflicting labels"
        )
    frame["_source_row_count"] = frame.groupby(dedup_key, dropna=False)["Response"].transform("size")
    frame = frame.drop_duplicates(dedup_key, keep="first").reset_index(drop=True)

    labels = frame["Response"].map(_label_from_response)
    specimen = frame["Dataset"].astype(str).str.strip().str.lower()
    out = pd.DataFrame(
        {
            "candidate_id": "cd8_multimer_2025:" + frame["_source_row"].astype(str),
            "source_dataset": "cd8_multimer_2025",
            "study_id": "cd8_multimer_2025",
            "patient_id": "cd8_multimer_2025:" + frame["Patient ID"].astype(str),
            "hla_allele": frame["HLA"].astype(str),
            "mutant_peptide": frame["MUT epitope"].astype(str),
            "wildtype_peptide": frame["WT epitope"].astype(str),
            "label": labels,
            "label_weight": 1.0,
            "assay_type": "phla_multimer_" + specimen,
            "source_patient_id": frame["Patient ID"].astype(str),
            "source_patient_alias": frame.get("Alt. ID", "").astype(str),
            "specimen_source": frame["Dataset"].astype(str),
            "tumor_type": frame["Tumor type"].astype(str),
            "gene_symbol": frame.get("SYMBOL", "").astype(str),
            "ensembl_gene_id": frame.get("ENSG", "").astype(str),
            "genome_assembly": frame.get("Genome assembly", "").astype(str),
            "source_duplicate_count": frame["_source_row_count"].astype(str),
        }
    )

    numeric_map = {
        "Binding affinity (%Rank score)": "netmhcpan_binding_percentile_rank",
        "RNA expression (TPM)": "expression_tpm",
        "Proteasomal processing score": "proteasomal_processing_score",
        "EL (%Rank score)": "netmhcpan_el_percentile_rank",
        "RF classifier score": "source_rf_presentation_score",
        "Agretopicity": "agretopicity_ratio",
        "Foreignness score": "foreignness_score",
        "Dissimilarity": "self_dissimilarity_score",
        "TMB": "tumor_mutational_burden",
    }
    for source, target in numeric_map.items():
        if source in frame.columns:
            out[target] = pd.to_numeric(frame[source], errors="coerce")

    if "netmhcpan_binding_percentile_rank" in out.columns:
        rank = out["netmhcpan_binding_percentile_rank"].clip(lower=0, upper=100)
        out["netmhcpan_binding_score"] = 1.0 - rank / 100.0
    if "netmhcpan_el_percentile_rank" in out.columns:
        rank = out["netmhcpan_el_percentile_rank"].clip(lower=0, upper=100)
        out["netmhcpan_el_score"] = 1.0 - rank / 100.0
    if "expression_tpm" in out.columns:
        out["log_expression_tpm"] = np.log1p(out["expression_tpm"].clip(lower=0))
    if "agretopicity_ratio" in out.columns:
        out["agretopicity_score"] = -np.log10(out["agretopicity_ratio"].clip(lower=1e-12))

    out = add_contrastive_features(out)
    out = add_normalized_columns(out)
    report = validate_schema(out)
    if not report.ok:
        raise ValueError(f"Normalized CD8 multimer table failed canonical schema validation: {report}")
    return out


def normalize_improve_cv(
    path: str | Path,
    *,
    zip_member: str | None = None,
) -> pd.DataFrame:
    """Normalize IMPROVE's official patient-disjoint cross-validation matrix."""
    table_path = Path(path)
    if table_path.suffix.lower() == ".zip" and zip_member is None:
        zip_member = (
            "data/03_data_for_CV/IMPROVE/"
            "03_3_final_peptide_features_Partition_for_CV.txt"
        )
    frame = read_table(table_path, zip_member=zip_member)
    required = {
        "Patient",
        "HLA_allele",
        "Norm_peptide",
        "Mut_peptide",
        "response",
        "cohort",
        "Partition",
    }
    missing = required - set(frame.columns)
    if missing:
        raise ValueError(f"IMPROVE CV table missing required columns: {sorted(missing)}")

    labels = frame["response"].map(_label_from_response)
    if (labels == "unknown").any():
        raise ValueError("IMPROVE CV table contains unrecognized response labels")

    out = pd.DataFrame(
        {
            "candidate_id": [f"improve:{idx}" for idx in range(len(frame))],
            "source_dataset": "improve",
            "study_id": "improve_" + frame["cohort"].astype(str),
            "patient_id": "improve:" + frame["Patient"].astype(str),
            "hla_allele": frame["HLA_allele"].astype(str),
            "mutant_peptide": frame["Mut_peptide"].astype(str),
            "wildtype_peptide": frame["Norm_peptide"].astype(str),
            "label": labels,
            "label_weight": 1.0,
            "assay_type": "validated_tcell_response",
            "official_partition": frame["Partition"].astype(str),
            "tumor_type": frame["cohort"].astype(str),
            "gene_symbol": frame.get("Gene_Symbol", "").astype(str),
        }
    )

    numeric_map = {
        "RankEL_4.1": "netmhcpan_el_percentile_rank",
        "RankBA_4.1": "netmhcpan_binding_percentile_rank",
        "RankEL_wt_4.1": "wildtype_netmhcpan_el_percentile_rank",
        "Stability": "binding_stability_hours",
        "Prime": "prime_source_score",
        "DAI_4.1": "differential_agretopicity_index",
        "Expression": "expression_tpm",
        "CelPrev": "cellular_prevalence",
        "PrioScore": "source_priority_score",
        "SelfSim": "self_similarity_score",
        "Foreigness": "foreignness_score",
        "HLAexp": "hla_expression",
        "CYT": "cytolytic_activity",
        "NetMHCExp": "netmhc_expression_score",
        "rna_af": "rna_variant_allele_fraction",
        "ValMutRNACoef": "validated_mutant_rna_coefficient",
        "HydroAll": "source_hydrophobicity_all",
        "HydroCore": "source_hydrophobicity_tcr_core",
    }
    for source, target in numeric_map.items():
        if source in frame.columns:
            out[target] = pd.to_numeric(frame[source], errors="coerce")

    if "netmhcpan_el_percentile_rank" in out.columns:
        rank = out["netmhcpan_el_percentile_rank"].clip(lower=0, upper=100)
        out["netmhcpan_el_score"] = 1.0 - rank / 100.0
    if "netmhcpan_binding_percentile_rank" in out.columns:
        rank = out["netmhcpan_binding_percentile_rank"].clip(lower=0, upper=100)
        out["netmhcpan_binding_score"] = 1.0 - rank / 100.0
    if "wildtype_netmhcpan_el_percentile_rank" in out.columns:
        rank = out["wildtype_netmhcpan_el_percentile_rank"].clip(lower=0, upper=100)
        out["wildtype_netmhcpan_el_score"] = 1.0 - rank / 100.0
    if "expression_tpm" in out.columns:
        out["log_expression_tpm"] = np.log1p(out["expression_tpm"].clip(lower=0))

    out = add_contrastive_features(out)
    out = add_normalized_columns(out)
    report = validate_schema(out)
    if not report.ok:
        raise ValueError(f"Normalized IMPROVE table failed canonical schema validation: {report}")
    return out


def write_normalized(frame: pd.DataFrame, output_path: str | Path) -> Path:
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(out, index=False)
    return out
